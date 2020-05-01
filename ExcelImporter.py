import openpyxl
from openpyxl.styles import Font

# Una class ens permet definir un "Objecte", en aquest cas un proveidor
# els objectes ens deixen guardar les dade per cada proviedor
class Product:
    def __init__(self):
        self.providerName = ''
        self.productID = ''
        self.name = ''
        self.qtMax = 0.0
        self.qtMin = 0.0
        self.leanPercentage = 0.0
        self.price = 0.0

    def toString(self):
        print('**********************************')
        print('Product product data: ')
        print('Provider name: ' + self.providerName)
        print('Product ID: ' + self.productID)
        print('Product Name: ' + self.name)
        print('Maximum quantity: ' + str(self.qtMax))
        print('Minimum quantity: ' + str(self.qtMin))
        print('Lean percentage: ' + str(self.leanPercentage))
        print('Price: ' + str(self.price))
        print('**********************************')

# i el mateix per cada commanda
class Order:
    def __init__(self):
        self.orderNumber = ''
        self.productName = ''
        self.meatKg = 0
        self.leanPercentage = 0.0
        self.day = ' '

    def toString(self):
        print('**********************************')
        print('Order data: ')
        print('Order number: ' + self.orderNumber)
        print('Meat Kg: ' + str(self.meatKg))
        print('Lean percentage: ' + str(self.leanPercentage))
        print('Order day: ' + self.day)
        print('**********************************')





# Get the row and column index for a given cell content
def GetRowColFromTableName(sheet, content):
    indexes = []
    for rowidx in list(range(1, sheet.max_row + 1)):
        for colidx in list(range(1, sheet.max_column + 1)):
            if sheet.cell(row = rowidx, column=colidx).value == content:
                # We add 1 to row, since we are looking for the title and we want to start
                # reading the content
                indexes.append( rowidx+1)
                indexes.append( colidx )
                return indexes

def GetData(filepath):
    
    ## CONFIGURATION DATA ##
    STARTING_PROVIDER_COLUMN_NAME = 'Proveïdor'
    STARTING_ORDER_COLUMN_NAME = 'Numero de comanda'

    # Aqui simplement definim un Array, que es una llista de coses, en aquest cas, de proveidors i una altra de comandes
    products = []
    orders = []

    # Agafem el name de l'excel que volem llegir
    print('Parsing file ' + filepath + '...')
    # El llegim i definim la fulla
    file = openpyxl.load_workbook(filepath)
    sheet = file.get_sheet_by_name("Dad_Ent")

    # Lets get where the products and orders info starts
    productsIndex = GetRowColFromTableName(sheet, STARTING_PROVIDER_COLUMN_NAME)
    productsRow  = productsIndex[0] 
    productsCol = productsIndex[1]

    ordersIndex = GetRowColFromTableName(sheet, STARTING_ORDER_COLUMN_NAME)
    ordersRow = ordersIndex[0]
    ordersCol = ordersIndex[1]

    # Parse products data
    for rowidx in list(range(productsRow, sheet.max_row + 1)):
        # Creem un objecte proveidor on guardarem les dades
        tempProduct = Product()
        tempProduct.providerName = sheet.cell(row = rowidx, column=productsCol).value
        tempProduct.productID = sheet.cell(row = rowidx, column=productsCol+1).value
        tempProduct.name = sheet.cell(row = rowidx, column=productsCol+2).value
        tempProduct.qtMax = sheet.cell(row = rowidx, column=productsCol+3).value
        tempProduct.qtMin = sheet.cell(row = rowidx, column=productsCol+4).value
        tempProduct.leanPercentage = sheet.cell(row = rowidx, column=productsCol+5).value
        tempProduct.price = sheet.cell(row = rowidx, column=productsCol+6).value
        # Check if the new product was created successfully
        if tempProduct.name != None:
            products.append(tempProduct)

    # Parse orders data
    for rowidx in list(range(ordersRow, sheet.max_row + 1)):
        # Comencem a partir de la 4a fila que es on comencen les dades
        tempCommanda = Order()
        tempCommanda.orderNumber = sheet.cell(row = rowidx, column=ordersCol).value
        tempCommanda.productName = sheet.cell(row = rowidx, column=ordersCol+1).value
        tempCommanda.meatKg = sheet.cell(row = rowidx, column=ordersCol+2).value
        tempCommanda.leanPercentage = sheet.cell(row = rowidx, column=ordersCol+3).value
        tempCommanda.day = sheet.cell(row = rowidx, column=ordersCol+4).value
        if tempCommanda.orderNumber != None:
            orders.append(tempCommanda)

    return products, orders

def GetArgumentsFromName(name):
    result = {}
    result["varName"] = name.split("[")[0]
    result["varSupp"] = name.split("[")[1].split(",")[0]
    result["varPID"] = name.split("[")[1].split(",")[1]
    result["varDay"] = name.split("[")[1].split(",")[2].replace("]","")

    return result

def GetProductNameFromID(productList, productID):
    for p in productList:
        if p.productID == productID:
            return p.name

def ExportDataToExcel(filename, products, suppliers, days, modelData ):
    
    print('Exporting data to ' + filename)
    # 1. Load excel file and the sheet we will work on
    file = openpyxl.load_workbook(filename)
    #sheet = file["Dad_Sort"]
    sheet = file.create_sheet(title="Dad_Sort2")
    sheet.cell(row=13, column=1).value = "DILLUNS"


    Amount = []
    Use = []
    Buy = []
    Stock = []

    # Distribute data by model variable
    for dataValue in modelData:
        if dataValue.X > 0:
            if dataValue.Varname.startswith("Amount"):
                Amount.append(dataValue)
            elif dataValue.Varname.startswith("Use"):
                Use.append(dataValue)
            elif dataValue.Varname.startswith("Buy"):
                Buy.append(dataValue)
            elif dataValue.Varname.startswith("Stock"):
                Stock.append(dataValue)
                

    # 2. We will iterate over every day, every supplier and every 
    currentDay = 0
    for day in days:
        # Get the index where the name of the day is written
        dayCellIndex = GetRowColFromTableName(sheet, day.upper())

        # Get the starting index to write data
        # Starting stock
        startRow = dayCellIndex[0] + 1
        startColumn = dayCellIndex[1] + 1
        # Get starting coordinates to 
        suppliersStartRow = dayCellIndex[0]
        suppliersStartColumn = startColumn + 1

        # Write default table values
        sheet.cell(row=dayCellIndex[0], column=(dayCellIndex[1]+1)).value = "Estoc Inicial"
        sheet.cell(row=dayCellIndex[0], column=(dayCellIndex[1]+1)).font = Font(bold=True)

        sheet.cell(row=startRow - 1, column=startColumn).value = "ID"
        sheet.cell(row=startRow - 1, column=startColumn+1).value = "NOM"

        # Now we are writing the values for every provider
        # for amountValue in Amount:
        supplierList = []
        for amountVar in Amount:
            varInfo = GetArgumentsFromName(amountVar.varName)

            # If the supplier does not have a column yet, write it
            if varInfo["varSupp"] not in supplierList:
                supplierList.append(varInfo["varSupp"])
                sheet.cell(row = suppliersStartRow, column= suppliersStartColumn+len(supplierList)).value = varInfo["varSupp"]
                print("Day: " + day + " - New supplier: " + varInfo["varSupp"] + " on coord: " + str(suppliersStartRow) + "," + str(suppliersStartColumn+len(supplierList)))

            # Write the product info
            if amountVar.X is not 0:
                sheet.cell(row = startRow, column=startColumn).value = varInfo["varPID"]
                sheet.cell(row = startRow, column=startColumn+1).value = GetProductNameFromID(products, varInfo["varPID"])
                sheet.cell(row = startRow, column=suppliersStartColumn+supplierList.index(varInfo["varSupp"])+1).value = amountVar.X
                
                startRow = startRow + 1

        # Reset the column index 
        currentDay = currentDay + 1
        if currentDay == len(days):
            currentDay = len(days) - 1
        sheet.cell(row = suppliersStartRow - 1, column=suppliersStartColumn+len(supplierList)+1).value = days[currentDay].upper()
        print("writing day: " + days[currentDay].upper() + " on coord: " + str(suppliersStartRow - 1) +" "+str(suppliersStartColumn+len(supplierList)))
        startColumn = dayCellIndex[1] + 1

    file.save(filename)
    print('Exporting done')

    